"""
FFB CCTV Grid Scanner
=====================
Overhead CCTV scans a ramp floor where FFB are laid flat (not stacked).
Frame is divided into a grid — ONE bunch per cell.
Each cell is classified, sized, and weighted.
Results exported to Excel with:
  - Sheet 1: Summary dashboard with KPIs
  - Sheet 2: Visual grid map (cell images embedded in grid layout)
  - Sheet 3: Row-by-row detail table with per-cell image

Usage:
    python ffb_grid_scanner.py --demo
    python ffb_grid_scanner.py --source ramp.mp4
    python ffb_grid_scanner.py --source photo.jpg
    python ffb_grid_scanner.py --source 0        # webcam / RTSP URL
"""

import argparse, io, os, math, random, sys
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field

import cv2
import numpy as np
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────

CFG = {
    "grid_rows":          5,       # rows in scan grid
    "grid_cols":          8,       # columns
    "scan_width_m":       6.0,     # real-world camera view width  (metres)
    "scan_height_m":      4.0,     # real-world camera view height
    "price_per_tonne_rm": 850.0,   # RM/tonne FFB price
    "output_dir":         "./ffb_scan_output",
    "excel_name":         "FFB_Scan_Report.xlsx",
    "cell_thumb_px":      (140, 110),  # saved thumbnail per cell
    "model_path":         "ffb_model.onnx",
}

Path(CFG["output_dir"]).mkdir(exist_ok=True)
CELL_DIR = Path(CFG["output_dir"]) / "cells"
CELL_DIR.mkdir(exist_ok=True)

# ──────────────────────────────────────────────
# MPOB GRADING CONSTANTS
# ──────────────────────────────────────────────

CLASSES = ["ripe", "underripe", "unripe", "rotten", "empty", "none"]

CLASS_LABEL = {
    "ripe":      "Ripe",
    "underripe": "Underripe",
    "unripe":    "Unripe",
    "rotten":    "Rotten",
    "empty":     "Empty",
    "none":      "—",
}

# BGR colour for OpenCV overlay
CLASS_BGR = {
    "ripe":      (40,  190,  60),
    "underripe": (30,  165, 255),
    "unripe":    (30,   30, 210),
    "rotten":    (160,  20, 180),
    "empty":     (100, 100, 100),
    "none":      (220, 220, 220),
}

# Excel cell background (hex, no #)
CLASS_HEX = {
    "ripe":      "C8E6C9",
    "underripe": "FFF9C4",
    "unripe":    "FFCDD2",
    "rotten":    "E1BEE7",
    "empty":     "ECEFF1",
    "none":      "F5F5F5",
}

# Oil extraction penalty per 1 % of that class in consignment (MPOB Tables III–VI)
OIL_PEN = {"ripe": 0.0, "underripe": 0.030, "unripe": 0.120,
           "rotten": 0.120, "empty": 0.100, "none": 0.0}

BASIC_OER  = 20.5   # % — Peninsular, Tenera DxP, 4-18 yrs (MPOB Table I)
BASIC_KER  = 5.5

# MPOB Table II bunch weight classes
def weight_class(kg: float) -> str:
    if kg < 5:    return "< 5 kg"
    if kg < 7:    return "5 – 7 kg"
    if kg < 10:   return "7 – 10 kg"
    if kg < 25:   return "10 – 25 kg"
    return "> 25 kg"

# ──────────────────────────────────────────────
# BUNCH DATA CLASS
# ──────────────────────────────────────────────

@dataclass
class BunchResult:
    grid_id:    str          # e.g. "R2C4"
    row:        int
    col:        int
    cls:        str
    conf:       float        # 0-1
    area_frac:  float        # bunch area / cell area  (0-1)
    weight_kg:  float
    thumb_path: str          # saved cell image
    frame_no:   int

    # derived
    label:       str = field(init=False)
    wt_class:    str = field(init=False)
    oil_pen:     float = field(init=False)
    graded_oer:  float = field(init=False)
    graded_ker:  float = field(init=False)
    price_rm:    float = field(init=False)

    def __post_init__(self):
        self.label      = CLASS_LABEL[self.cls]
        self.wt_class   = weight_class(self.weight_kg) if self.cls != "none" else "—"
        pct             = self.conf * 100
        self.oil_pen    = round(OIL_PEN[self.cls] * pct, 4)
        self.graded_oer = round(max(0.0, BASIC_OER - self.oil_pen), 2)
        self.graded_ker = BASIC_KER
        if self.cls not in ("none", "empty"):
            self.price_rm = round(
                self.weight_kg / 1000 * CFG["price_per_tonne_rm"]
                * (self.graded_oer / BASIC_OER), 2)
        else:
            self.price_rm = 0.0

# ──────────────────────────────────────────────
# WEIGHT ESTIMATION
# Bunch apparent area in image → real-world diameter → MPOB weight proxy
# ──────────────────────────────────────────────

def estimate_weight(area_frac: float, frame_w: int, frame_h: int) -> float:
    """
    area_frac: fraction of the full frame the bunch occupies.
    Converts pixel area → real-world area → diameter → weight.
    Weight range 2–35 kg mapped from diameter 0.20–0.60 m.
    """
    real_area_m2 = area_frac * CFG["scan_width_m"] * CFG["scan_height_m"]
    diameter_m   = 2 * math.sqrt(max(real_area_m2, 1e-6) / math.pi)
    diameter_m   = max(0.20, min(0.60, diameter_m))
    t            = (diameter_m - 0.20) / 0.40      # 0 → 1
    kg           = 2 + t * 33                       # 2 kg … 35 kg
    return round(kg, 1)

# ──────────────────────────────────────────────
# CLASSIFIER  (ONNX if available, else colour heuristic)
# ──────────────────────────────────────────────

MEAN = np.array([0.485, 0.456, 0.406], dtype=np.float32)
STD  = np.array([0.229, 0.224, 0.225], dtype=np.float32)

def load_model(path: str):
    try:
        import onnxruntime as ort
        if Path(path).exists():
            sess = ort.InferenceSession(path, providers=["CPUExecutionProvider"])
            print(f"✅  ONNX model loaded: {path}")
            return sess
    except ImportError:
        pass
    print("⚠️   No ONNX model — using colour-heuristic classifier")
    return None

def _softmax(x):
    e = np.exp(x - x.max()); return e / e.sum()

def classify(crop: np.ndarray, model) -> tuple[str, float]:
    if model is not None:
        rgb  = cv2.cvtColor(cv2.resize(crop, (224, 224)), cv2.COLOR_BGR2RGB)
        arr  = ((rgb.astype(np.float32)/255 - MEAN) / STD).transpose(2,0,1)[None]
        probs = _softmax(model.run(None, {model.get_inputs()[0].name: arr})[0][0])
        idx   = int(np.argmax(probs))
        return CLASSES[idx], float(probs[idx])
    return _heuristic(crop)

def _heuristic(bgr: np.ndarray) -> tuple[str, float]:
    """Fast colour heuristic for demo / no-model fallback."""
    if bgr is None or bgr.size == 0:
        return "none", 1.0
    hsv = cv2.cvtColor(bgr, cv2.COLOR_BGR2HSV).astype(np.float32)
    h, s, v = hsv[...,0], hsv[...,1], hsv[...,2]
    mv, ms = float(v.mean()), float(s.mean())

    if mv > 200 and ms < 25:          return "none",      0.90
    if mv < 55:                        return "unripe",    0.72
    if mv < 85  and ms < 45:          return "rotten",    0.68
    orange = ((h < 22) | (h > 158)) & (s > 55)
    if orange.mean() > 0.30:          return "ripe",      min(0.95, 0.55 + float(orange.mean()))
    purple = (h > 118) & (h < 162) & (s > 45)
    if purple.mean() > 0.18:          return "underripe", 0.66
    if ms < 30:                        return "empty",     0.62
    return "ripe", 0.52

# ──────────────────────────────────────────────
# BUNCH AREA DETECTION  (segment bunch from ramp floor)
# ──────────────────────────────────────────────

def bunch_area_fraction(cell: np.ndarray) -> float:
    """Return fraction of cell area occupied by the bunch (0-1)."""
    hsv  = cv2.cvtColor(cell, cv2.COLOR_BGR2HSV)
    mask = cv2.inRange(hsv, np.array([0, 18, 25]), np.array([180, 255, 215]))
    k    = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 7))
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, k)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN,  k)
    total = cell.shape[0] * cell.shape[1]
    return float(np.count_nonzero(mask)) / max(total, 1)

# ──────────────────────────────────────────────
# GRID SCAN
# ──────────────────────────────────────────────

def scan_frame(frame: np.ndarray, model, frame_no: int) -> list[BunchResult]:
    fh, fw  = frame.shape[:2]
    nrows, ncols = CFG["grid_rows"], CFG["grid_cols"]
    cell_h  = fh // nrows
    cell_w  = fw // ncols
    pad_h   = max(2, int(cell_h * 0.03))
    pad_w   = max(2, int(cell_w * 0.03))

    results = []
    for r in range(nrows):
        for c in range(ncols):
            y0, y1 = r*cell_h + pad_h,  (r+1)*cell_h - pad_h
            x0, x1 = c*cell_w + pad_w,  (c+1)*cell_w - pad_w
            crop    = frame[y0:y1, x0:x1]

            cls, conf  = classify(crop, model)
            af         = bunch_area_fraction(crop)
            # Scale cell area fraction up to full-frame fraction
            cell_frac  = (cell_h * cell_w) / (fh * fw)
            full_frac  = af * cell_frac
            weight     = estimate_weight(full_frac, fw, fh) if cls != "none" else 0.0
            grid_id    = f"R{r+1}C{c+1}"

            # Save thumbnail
            thumb_path = str(CELL_DIR / f"{grid_id}_f{frame_no:04d}.jpg")
            cv2.imwrite(thumb_path, cv2.resize(crop, CFG["cell_thumb_px"]))

            results.append(BunchResult(
                grid_id=grid_id, row=r+1, col=c+1,
                cls=cls, conf=conf,
                area_frac=full_frac, weight_kg=weight,
                thumb_path=thumb_path, frame_no=frame_no,
            ))
    return results

# ──────────────────────────────────────────────
# ANNOTATE FRAME  (live overlay for display)
# ──────────────────────────────────────────────

def annotate(frame: np.ndarray, results: list[BunchResult]) -> np.ndarray:
    out  = frame.copy()
    fh, fw = out.shape[:2]
    nrows, ncols = CFG["grid_rows"], CFG["grid_cols"]
    ch, cw = fh // nrows, fw // ncols

    for res in results:
        r, c   = res.row-1, res.col-1
        x0, y0 = c*cw, r*ch
        x1, y1 = x0+cw, y0+ch
        color  = CLASS_BGR[res.cls]

        # Semi-transparent fill
        ov = out.copy()
        cv2.rectangle(ov, (x0,y0), (x1,y1), color, -1)
        cv2.addWeighted(ov, 0.22, out, 0.78, 0, out)

        # Border — thicker for high-confidence ripe
        thick = 3 if (res.cls == "ripe" and res.conf > 0.7) else 1
        cv2.rectangle(out, (x0,y0), (x1,y1), color, thick)

        # Grid ID top-left
        cv2.putText(out, res.grid_id, (x0+4, y0+14),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.38, (255,255,255), 1, cv2.LINE_AA)
        if res.cls != "none":
            # Class + confidence
            cv2.putText(out, f"{res.label[:5]} {res.conf*100:.0f}%",
                        (x0+4, y0+28), cv2.FONT_HERSHEY_SIMPLEX, 0.33, color, 1, cv2.LINE_AA)
            # Weight
            cv2.putText(out, f"{res.weight_kg:.1f}kg",
                        (x0+4, y0+42), cv2.FONT_HERSHEY_SIMPLEX, 0.33, (230,230,230), 1, cv2.LINE_AA)
            # Price
            cv2.putText(out, f"RM{res.price_rm:.2f}",
                        (x0+4, y0+56), cv2.FONT_HERSHEY_SIMPLEX, 0.30, (200,240,200), 1, cv2.LINE_AA)

    # Status bar
    bunches = [r for r in results if r.cls != "none"]
    total_w = sum(r.weight_kg for r in bunches)
    total_p = sum(r.price_rm  for r in bunches)
    bar = (f"  Bunches: {len(bunches)}/{len(results)}   "
           f"Weight: {total_w:.1f} kg   "
           f"Value: RM {total_p:.2f}   "
           f"{datetime.now().strftime('%H:%M:%S')}")
    cv2.rectangle(out, (0, fh-24), (fw, fh), (20,20,20), -1)
    cv2.putText(out, bar, (6, fh-7),
                cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0,230,0), 1, cv2.LINE_AA)
    return out

# ──────────────────────────────────────────────
# DEMO FRAME GENERATOR
# Simulates an overhead shot of bunches laid flat on a ramp
# ──────────────────────────────────────────────

DEMO_LAYOUT = [
    ["ripe","ripe","ripe","underripe","ripe","ripe","ripe","ripe"],
    ["ripe","unripe","ripe","ripe","rotten","ripe","underripe","ripe"],
    ["underripe","ripe","ripe","ripe","ripe","empty","ripe","ripe"],
    ["ripe","ripe","none","ripe","ripe","ripe","ripe","unripe"],
    ["ripe","ripe","ripe","rotten","ripe","ripe","none","ripe"],
]

BUNCH_PALETTE = {
    "ripe":      [(30,140,220),(40,155,235),(25,125,210),(35,145,228)],
    "underripe": [(75,95,200),(95,85,190),(65,105,210),(85,90,195)],
    "unripe":    [(35,35,85), (30,30,75), (40,40,95), (32,32,80)],
    "rotten":    [(55,18,75), (48,22,68), (60,15,82), (52,20,72)],
    "empty":     [(80,95,105),(85,100,112),(90,105,118),(78,92,102)],
}

def make_demo_frame(width=1280, height=800) -> np.ndarray:
    """Render a photorealistic-enough ramp floor with FFB laid flat."""
    # Ramp concrete floor
    rng = np.random.default_rng(42)
    base = rng.integers(185, 205, (height, width, 3), dtype=np.uint8)
    # Dirt / texture noise
    noise = rng.integers(0, 18, (height, width), dtype=np.uint8)
    base  = np.clip(base.astype(np.int16) - noise[:,:,None], 0, 255).astype(np.uint8)
    frame = base

    nrows, ncols = CFG["grid_rows"], CFG["grid_cols"]
    ch = height // nrows
    cw = width  // ncols

    for r in range(nrows):
        for c in range(ncols):
            cls = DEMO_LAYOUT[r % len(DEMO_LAYOUT)][c % len(DEMO_LAYOUT[0])]
            if cls == "none":
                continue

            cx = int((c + 0.5) * cw)
            cy = int((r + 0.5) * ch)

            # Bunch ellipse radii (varied per cell)
            rx = int(cw * random.uniform(0.30, 0.42))
            ry = int(ch * random.uniform(0.30, 0.42))
            rot = random.randint(0, 180)

            base_color = random.choice(BUNCH_PALETTE[cls])

            # Draw main ellipse
            cv2.ellipse(frame, (cx,cy), (rx,ry), rot, 0, 360, base_color, -1)

            # Fruitlet texture (small circles)
            n_fruits = 60 if cls != "empty" else 15
            for _ in range(n_fruits):
                ang  = random.uniform(0, 2*math.pi)
                dist = random.uniform(0, 0.88)
                fx   = int(cx + dist*rx*math.cos(ang))
                fy   = int(cy + dist*ry*math.sin(ang))
                fr   = random.randint(3, 9)
                fc   = tuple(max(0, min(255, v + random.randint(-25,25)))
                             for v in base_color)
                cv2.circle(frame, (fx,fy), fr, fc, -1)

            # Shadow under bunch
            shadow = frame.copy()
            cv2.ellipse(shadow, (cx+4,cy+5), (rx,ry), rot, 0, 360, (40,40,40), -1)
            cv2.addWeighted(shadow, 0.12, frame, 0.88, 0, frame)
            cv2.ellipse(frame, (cx,cy), (rx,ry), rot, 0, 360, base_color, -1)
            # Re-draw fruitlets on top
            for _ in range(n_fruits):
                ang  = random.uniform(0, 2*math.pi)
                dist = random.uniform(0, 0.88)
                fx   = int(cx + dist*rx*math.cos(ang))
                fy   = int(cy + dist*ry*math.sin(ang))
                fr   = random.randint(3, 9)
                fc   = tuple(max(0, min(255, v + random.randint(-25,25)))
                             for v in base_color)
                cv2.circle(frame, (fx,fy), fr, fc, -1)

            # Peduncle stub
            px = int(cx + rx*0.7*math.cos(math.radians(rot+15)))
            py = int(cy + ry*0.7*math.sin(math.radians(rot+15)))
            cv2.line(frame, (cx,cy), (px,py), (30,60,30), 3)

    return frame

# ──────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────

def _s(style="thin"):
    return Side(style=style, color="BDBDBD")

BORDER  = Border(left=_s(), right=_s(), top=_s(), bottom=_s())
HFILL   = PatternFill("solid", fgColor="1B5E20")
HFONT   = Font(bold=True, color="FFFFFF", name="Arial", size=9)
BFONT   = Font(name="Arial", size=9)
TFONT   = Font(name="Arial", size=14, bold=True, color="1B5E20")
SMFONT  = Font(name="Arial", size=8, italic=True, color="616161")
CTR     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _w(ws, rc, val, font=None, fill=None, aln=None, bdr=None, fmt=None):
    c = ws[rc]
    c.value = val
    if font: c.font      = font
    if fill: c.fill      = fill
    if aln:  c.alignment = aln
    if bdr:  c.border    = bdr
    if fmt:  c.number_format = fmt


def export_excel(results: list[BunchResult],
                 annotated_frame: np.ndarray,
                 meta: dict) -> str:

    wb = Workbook()
    _sheet_summary(wb, results, annotated_frame, meta)
    _sheet_grid(wb, results)
    _sheet_detail(wb, results)

    out = str(Path(CFG["output_dir"]) / CFG["excel_name"])
    wb.save(out)
    return out

# ── Sheet 1: Summary ─────────────────────────────────────────

def _sheet_summary(wb, results, ann_frame, meta):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 14

    # Title
    ws.merge_cells("A1:H1")
    _w(ws,"A1","🌴  MPOB Oil Palm FFB — CCTV Grid Scan Report", TFONT, aln=CTR)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    _w(ws,"A2",
       f"Scanned: {meta['ts']}   |   Source: {meta['source']}   |   "
       f"Grid: {CFG['grid_rows']}×{CFG['grid_cols']}   |   "
       f"Price: RM {CFG['price_per_tonne_rm']:.0f}/tonne",
       SMFONT, aln=LFT)
    ws.row_dimensions[2].height = 14

    # Annotated frame snapshot
    img_end_row = 3
    if ann_frame is not None:
        h, w = 220, 390
        thumb = cv2.cvtColor(cv2.resize(ann_frame, (w,h)), cv2.COLOR_BGR2RGB)
        buf   = io.BytesIO()
        Image.fromarray(thumb).save(buf, "PNG"); buf.seek(0)
        xl    = XLImage(buf); xl.width = w; xl.height = h
        ws.add_image(xl, "A4")
        # fix row heights to accommodate image
        for rr in range(4, 4 + 15):
            ws.row_dimensions[rr].height = h / 15
        img_end_row = 19

    # KPI row
    bunches  = [r for r in results if r.cls != "none"]
    total_w  = sum(r.weight_kg for r in bunches)
    total_p  = sum(r.price_rm  for r in bunches)
    ripe_n   = sum(1 for r in bunches if r.cls == "ripe")
    ripe_pct = ripe_n / max(len(bunches),1) * 100

    kpi_row = img_end_row + 1
    kpis = [
        ("A", "Total Bunches",   len(bunches)),
        ("C", "Total Weight (kg)", total_w),
        ("E", "Est. Value (RM)",  total_p),
        ("G", "Ripe %",           ripe_pct),
    ]
    fmts = ["0", "#,##0.0", "\"RM\"#,##0.00", "0.0\"%\""]
    for (col, lbl, val), fmt in zip(kpis, fmts):
        c2 = chr(ord(col)+1)
        ws.merge_cells(f"{col}{kpi_row}:{c2}{kpi_row}")
        ws.merge_cells(f"{col}{kpi_row+1}:{c2}{kpi_row+1}")
        _w(ws, f"{col}{kpi_row}", lbl, SMFONT, aln=CTR)
        cc = ws[f"{col}{kpi_row+1}"]
        cc.value = val; cc.number_format = fmt
        cc.font  = Font(name="Arial", size=20, bold=True, color="1B5E20")
        cc.alignment = CTR
    ws.row_dimensions[kpi_row].height   = 14
    ws.row_dimensions[kpi_row+1].height = 34

    # Breakdown table
    hdr_row = kpi_row + 3
    hdrs = ["Class","Bunches","% of Load","Weight (kg)",
            "Value (RM)","Basic OER %","Oil Penalty %","Graded OER %"]
    for i, h in enumerate(hdrs):
        _w(ws, f"{get_column_letter(i+1)}{hdr_row}",
           h, HFONT, HFILL, CTR, BORDER)
    ws.row_dimensions[hdr_row].height = 18

    cls_list = [c for c in CLASSES if c != "none"]
    for i, cls in enumerate(cls_list):
        rr   = hdr_row + 1 + i
        grp  = [r for r in results if r.cls == cls]
        cnt  = len(grp)
        wt   = sum(r.weight_kg for r in grp)
        val  = sum(r.price_rm  for r in grp)
        pct  = cnt / max(len(bunches),1)
        fill = PatternFill("solid", fgColor=CLASS_HEX[cls])
        row_vals = [CLASS_LABEL[cls], cnt, pct, wt, val,
                    BASIC_OER, OIL_PEN[cls], BASIC_OER - OIL_PEN[cls]*pct*100]
        row_fmts = ["@","0","0.0%","#,##0.0","\"RM\"#,##0.00","0.0","0.000","0.00"]
        for j,(v,f) in enumerate(zip(row_vals, row_fmts)):
            _w(ws, f"{get_column_letter(j+1)}{rr}", v, BFONT, fill, CTR, BORDER, f)
        ws.row_dimensions[rr].height = 15

    # Totals
    tot_row = hdr_row + 1 + len(cls_list)
    tfill   = PatternFill("solid", fgColor="E8F5E9")
    tfont   = Font(name="Arial", size=9, bold=True, color="1B5E20")
    tot_vals = ["TOTAL", len(bunches), 1.0, total_w, total_p, BASIC_OER, "", ""]
    tot_fmts = ["@","0","0%","#,##0.0","\"RM\"#,##0.00","0.0","",""]
    for j,(v,f) in enumerate(zip(tot_vals, tot_fmts)):
        _w(ws, f"{get_column_letter(j+1)}{tot_row}", v, tfont, tfill, CTR, BORDER, f)
    ws.row_dimensions[tot_row].height = 16

    # MPOB advisory note
    note_row = tot_row + 2
    ws.merge_cells(f"A{note_row}:H{note_row}")
    _w(ws, f"A{note_row}",
       "MPOB Quality Limits (Sec 5.2.3): Ripe ≥90%  Underripe ≤10%  Unripe/Rotten/Empty = 0%   "
       "Reject load if Empty >20% or Dirty >30% of consignment.",
       Font(name="Arial", size=8, italic=True, color="B71C1C"),
       PatternFill("solid", fgColor="FFF9C4"), LFT)
    ws.row_dimensions[note_row].height = 14

    disc_row = note_row + 1
    ws.merge_cells(f"A{disc_row}:H{disc_row}")
    _w(ws, f"A{disc_row}",
       "⚠  Prototype AI system. Results must be verified by a qualified MPOB-licensed grading officer (MPOB Manual Sec 3.2).",
       Font(name="Arial", size=8, italic=True, color="777777"),
       aln=LFT)

# ── Sheet 2: Grid Map — cell images embedded in grid layout ──

def _sheet_grid(wb, results):
    ws = wb.create_sheet("Grid Map")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:Z1")
    _w(ws,"A1","FFB Grid Map — Each cell shows bunch photo, class, weight and value",
       TFONT, aln=CTR)
    ws.row_dimensions[1].height = 26

    nrows, ncols = CFG["grid_rows"], CFG["grid_cols"]
    IMG_W, IMG_H = CFG["cell_thumb_px"]

    # Each grid cell = COLS_PER columns x ROWS_PER rows in Excel
    COLS_PER = 2
    ROWS_PER = 10
    START_R  = 3
    START_C  = 2   # col B onwards (col A = row labels)

    # Column headers
    for c in range(ncols):
        ec  = get_column_letter(START_C + c*COLS_PER)
        ec2 = get_column_letter(START_C + c*COLS_PER + 1)
        ws.merge_cells(f"{ec}2:{ec2}2")
        _w(ws, f"{ec}2", f"Col {c+1}", HFONT, HFILL, CTR)
        ws.column_dimensions[ec].width  = IMG_W / 7.5
        ws.column_dimensions[ec2].width = IMG_W / 7.5
    ws.column_dimensions["A"].width = 5

    rmap = {(r.row, r.col): r for r in results}

    for r in range(nrows):
        base = START_R + r * ROWS_PER
        # Row label
        ws.merge_cells(f"A{base}:A{base+ROWS_PER-1}")
        _w(ws, f"A{base}", f"Row {r+1}", HFONT, HFILL,
           Alignment(horizontal="center", vertical="center", text_rotation=90))
        for sub in range(ROWS_PER):
            ws.row_dimensions[base+sub].height = IMG_H / ROWS_PER * 1.1

        for c in range(ncols):
            res  = rmap.get((r+1, c+1))
            if not res: continue
            fill = PatternFill("solid", fgColor=CLASS_HEX[res.cls])
            ec   = get_column_letter(START_C + c*COLS_PER)
            ec2  = get_column_letter(START_C + c*COLS_PER + 1)

            # Image rows — do NOT merge, just anchor image
            img_row = base
            if Path(res.thumb_path).exists():
                xi = XLImage(res.thumb_path)
                xi.width = IMG_W; xi.height = IMG_H - 28
                ws.add_image(xi, f"{ec}{img_row}")

            # Colour the image rows
            for sub in range(ROWS_PER - 2):
                for col_off in range(COLS_PER):
                    cell = ws.cell(row=base+sub, column=START_C + c*COLS_PER + col_off)
                    cell.fill   = fill
                    cell.border = BORDER

            # Caption rows at bottom (last 2 rows) — write BEFORE merging
            cap_row = base + ROWS_PER - 2
            cap_cell = ws.cell(row=cap_row, column=START_C + c*COLS_PER)
            caption = f"{res.grid_id}  {res.label}\n{res.weight_kg:.1f} kg | RM {res.price_rm:.2f}\nConf: {res.conf*100:.0f}%"
            cap_cell.font      = Font(name="Arial", size=7, bold=True)
            cap_cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
            cap_cell.fill   = fill
            cap_cell.border = BORDER
            ws.merge_cells(f"{ec}{cap_row}:{ec2}{cap_row+1}")


def _sheet_detail(wb, results):
    ws = wb.create_sheet("Cell Detail")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:P1")
    _w(ws,"A1","Cell-by-Cell Detail", TFONT, aln=CTR)
    ws.row_dimensions[1].height = 22

    IMG_W, IMG_H = CFG["cell_thumb_px"]
    ROW_H = IMG_H * 0.75

    headers = [
        "Grid ID","Row","Col","Class","Confidence %",
        "MPOB Weight Class","Est. Weight (kg)","Est. Price (RM)",
        "Basic OER %","Oil Penalty %","Graded OER %","Graded KER %",
        "Area (% frame)","Frame No.","Scan Time","Cell Photo"
    ]
    for i,h in enumerate(headers):
        _w(ws, f"{get_column_letter(i+1)}2", h, HFONT, HFILL, CTR, BORDER)
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for i, res in enumerate(results):
        rr   = i + 3
        fill = PatternFill("solid", fgColor=CLASS_HEX[res.cls])
        vals = [
            res.grid_id, res.row, res.col, res.label,
            res.conf*100, res.wt_class, res.weight_kg, res.price_rm,
            BASIC_OER, res.oil_pen, res.graded_oer, res.graded_ker,
            res.area_frac*100, res.frame_no, now_str, ""
        ]
        fmts = [
            "@","0","0","@","0.0",
            "@","#,##0.0","\"RM\"#,##0.00",
            "0.0","0.0000","0.00","0.0",
            "0.00","0","@","@"
        ]
        for j,(v,f) in enumerate(zip(vals, fmts)):
            _w(ws, f"{get_column_letter(j+1)}{rr}", v, BFONT, fill, CTR, BORDER, f)

        # Embed cell photo in last column
        photo_col = get_column_letter(len(headers))
        if Path(res.thumb_path).exists():
            xi = XLImage(res.thumb_path)
            xi.width = IMG_W; xi.height = IMG_H
            ws.add_image(xi, f"{photo_col}{rr}")
        ws.row_dimensions[rr].height = ROW_H

    # Column widths
    widths = [8,5,5,12,12,14,14,14,10,11,11,10,12,9,17,18]
    for i,w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source",     default=None, help="Video/image path or camera index")
    ap.add_argument("--demo",       action="store_true")
    ap.add_argument("--grid-rows",  type=int, default=CFG["grid_rows"])
    ap.add_argument("--grid-cols",  type=int, default=CFG["grid_cols"])
    ap.add_argument("--price",      type=float, default=CFG["price_per_tonne_rm"])
    ap.add_argument("--no-display", action="store_true")
    args = ap.parse_args()

    CFG["grid_rows"]          = args.grid_rows
    CFG["grid_cols"]          = args.grid_cols
    CFG["price_per_tonne_rm"] = args.price

    model   = load_model(CFG["model_path"])
    results = []
    ann     = None
    source  = args.source or "demo"

    # ── Demo ──────────────────────────────────
    if args.demo or args.source is None:
        print("▶  Running demo scan…")
        frame   = make_demo_frame()
        results = scan_frame(frame, model, frame_no=0)
        ann     = annotate(frame, results)
        if not args.no_display:
            cv2.imshow("FFB Grid Scanner — DEMO (press any key)", ann)
            cv2.waitKey(0); cv2.destroyAllWindows()

    # ── Image file ────────────────────────────
    elif Path(args.source).suffix.lower() in (".jpg",".jpeg",".png",".bmp",".webp"):
        print(f"▶  Scanning image: {args.source}")
        frame = cv2.imread(args.source)
        if frame is None:
            sys.exit(f"Cannot read: {args.source}")
        results = scan_frame(frame, model, frame_no=0)
        ann     = annotate(frame, results)
        if not args.no_display:
            cv2.imshow("FFB Grid Scanner", ann)
            cv2.waitKey(0); cv2.destroyAllWindows()

    # ── Video / webcam ────────────────────────
    else:
        src = int(args.source) if args.source.isdigit() else args.source
        cap = cv2.VideoCapture(src)
        if not cap.isOpened():
            sys.exit(f"Cannot open: {args.source}")
        fps       = max(1, cap.get(cv2.CAP_PROP_FPS) or 25)
        every     = int(fps * 2)           # scan one frame every ~2 s
        frame_no  = 0
        best_ann  = None
        best_ripe = -1
        print(f"▶  Video scan — scanning every {every} frames. Press Q to stop.")
        while True:
            ok, frame = cap.read()
            if not ok: break
            frame_no += 1
            if frame_no % every == 0:
                r = scan_frame(frame, model, frame_no)
                results.extend(r)
                n_ripe = sum(1 for x in r if x.cls=="ripe")
                if n_ripe > best_ripe:
                    best_ripe = n_ripe
                    best_ann  = annotate(frame, r)
                print(f"  frame {frame_no}: {sum(1 for x in r if x.cls!='none')} bunches "
                      f"| {sum(x.weight_kg for x in r):.1f} kg")
            if not args.no_display:
                disp_r = [x for x in results if x.frame_no==frame_no]
                cv2.imshow("FFB Grid Scanner — Q to export", annotate(frame, disp_r) if disp_r else frame)
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
        cap.release()
        cv2.destroyAllWindows()
        ann = best_ann

    if not results:
        print("No results."); return

    # ── Print summary ─────────────────────────
    bunches = [r for r in results if r.cls!="none"]
    total_w = sum(r.weight_kg for r in bunches)
    total_p = sum(r.price_rm  for r in bunches)
    print(f"\n{'='*54}")
    print(f"  SCAN SUMMARY")
    print(f"{'='*54}")
    print(f"  Grid cells scanned : {len(results)}")
    print(f"  Bunches detected   : {len(bunches)}")
    print(f"  Total weight       : {total_w:.1f} kg")
    print(f"  Estimated value    : RM {total_p:.2f}")
    print(f"  {'Class':<14}  Count   % ")
    for cls in CLASSES:
        if cls=="none": continue
        n = sum(1 for r in bunches if r.cls==cls)
        if n: print(f"  {CLASS_LABEL[cls]:<14}  {n:>5}  {n/max(len(bunches),1)*100:>5.1f}%")
    print(f"{'='*54}")

    # ── Export Excel ──────────────────────────
    path = export_excel(
        results, ann,
        {"ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "source": source}
    )
    print(f"\n✅  Excel report  : {path}")
    print(f"   Cell images   : {CELL_DIR}/")
    return path

if __name__ == "__main__":
    main()
