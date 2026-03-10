"""
FFB Real-Time CCTV Scanner  — MPOB Oil Palm Grading System
===========================================================
Works directly with surveillance camera / CCTV feed OR video file.
Detects each fruit bunch using Hough circle transform, assigns a
grid ID, classifies ripeness, estimates weight & value, overlays
all labels live on screen, then exports full Excel report on exit.

Usage:
    python ffb_realtime_scanner.py                            # webcam 0
    python ffb_realtime_scanner.py --source video.mp4         # video file
    python ffb_realtime_scanner.py --source rtsp://...        # IP camera
    python ffb_realtime_scanner.py --source video.mp4 --export-frames

Controls (while window is open):
    S  – manual snapshot + save current detections to Excel
    R  – reset / clear all tracked bunches
    Q  – quit and auto-export final Excel report
    SPACE – pause / resume
"""

import argparse
import io
import math
import os
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

import cv2
import numpy as np
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────
# CONFIGURATION  — tune these to your ramp setup
# ──────────────────────────────────────────────────────────────────
CFG = {
    # Camera / scene
    "source":           0,              # 0=webcam, or path/URL
    "scene_width_m":    5.0,            # estimated real-world width of camera view
    "scene_height_m":   4.0,

    # Detection (Hough circles)
    "hough_dp":         1.2,
    "hough_min_dist":   38,             # min pixels between bunch centres
    "hough_p1":         50,
    "hough_p2":         22,
    "hough_min_r":      14,             # min bunch radius in pixels
    "hough_max_r":      75,             # max bunch radius in pixels

    # Grid (for grid-ID assignment)
    "grid_cols":        8,
    "grid_rows":        6,

    # Weight model (MPOB Table II proxy — diameter → weight)
    "weight_min_kg":    2.0,            # corresponds to min_radius bunch
    "weight_max_kg":    35.0,           # corresponds to max_radius bunch

    # Price
    "price_rm_per_tonne": 850.0,

    # Processing
    "process_every_n":  3,              # analyse every N frames (speed vs accuracy)
    "track_iou_thresh": 0.35,           # IoU threshold for bunch re-identification
    "stable_after_n":   2,              # frames a detection must persist to be "confirmed"

    # Output
    "output_dir":       "./ffb_output",
    "excel_name":       "FFB_Scan_Report.xlsx",
    "thumb_size":       (130, 100),     # cell image saved to Excel

    # Display
    "display_scale":    1.4,            # window zoom factor
    "show_debug":       False,
}

# ──────────────────────────────────────────────────────────────────
# MPOB CLASS DEFINITIONS
# ──────────────────────────────────────────────────────────────────
CLASSES = ["Ripe", "Underripe", "Unripe", "Rotten", "Empty"]

# BGR colours for overlay
C_BGR = {
    "Ripe":      (40,  200,  50),
    "Underripe": (20,  165, 255),
    "Unripe":    (30,   30, 200),
    "Rotten":    (160,  20, 180),
    "Empty":     (120, 120, 120),
}
# Excel cell fill hex
C_HEX = {
    "Ripe":      "C8E6C9",
    "Underripe": "FFF9C4",
    "Unripe":    "FFCDD2",
    "Rotten":    "E1BEE7",
    "Empty":     "ECEFF1",
}
# Oil extraction penalty per confidence-weighted % (MPOB Tables III–VI)
OIL_PEN = {
    "Ripe": 0.0, "Underripe": 0.030, "Unripe": 0.120,
    "Rotten": 0.120, "Empty": 0.100,
}
BASIC_OER = 20.5   # % — Peninsular, Tenera DxP, 4-18 yrs
BASIC_KER = 5.5    # %

def mpob_weight_class(kg: float) -> str:
    if kg < 5:  return "< 5 kg"
    if kg < 7:  return "5–7 kg"
    if kg < 10: return "7–10 kg"
    if kg < 25: return "10–25 kg"
    return "> 25 kg"

# ──────────────────────────────────────────────────────────────────
# BUNCH DATA
# ──────────────────────────────────────────────────────────────────
@dataclass
class Bunch:
    track_id:   int
    cx: int;  cy: int;  radius: int
    cls:        str  = "Ripe"
    conf:       float = 0.70
    weight_kg:  float = 10.0
    grid_id:    str   = ""
    stable_cnt: int   = 0          # frames seen
    confirmed:  bool  = False
    thumb:      Optional[np.ndarray] = field(default=None, repr=False)
    first_seen: str   = ""
    snap_frame: Optional[np.ndarray] = field(default=None, repr=False)

    # derived
    @property
    def wt_class(self) -> str:
        return mpob_weight_class(self.weight_kg)

    @property
    def oil_pen(self) -> float:
        return round(OIL_PEN[self.cls] * self.conf * 100, 4)

    @property
    def graded_oer(self) -> float:
        return round(max(0.0, BASIC_OER - self.oil_pen), 2)

    @property
    def price_rm(self) -> float:
        if self.cls in ("Empty",):
            return 0.0
        return round(self.weight_kg / 1000 * CFG["price_rm_per_tonne"]
                     * (self.graded_oer / BASIC_OER), 2)

# ──────────────────────────────────────────────────────────────────
# BUNCH DETECTION
# ──────────────────────────────────────────────────────────────────
def detect_bunches(frame: np.ndarray) -> list[tuple[int,int,int]]:
    """Return list of (cx, cy, radius) from Hough circle detection."""
    gray  = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    blur  = cv2.GaussianBlur(gray, (9, 9), 2)
    circles = cv2.HoughCircles(
        blur, cv2.HOUGH_GRADIENT,
        dp=CFG["hough_dp"], minDist=CFG["hough_min_dist"],
        param1=CFG["hough_p1"],  param2=CFG["hough_p2"],
        minRadius=CFG["hough_min_r"], maxRadius=CFG["hough_max_r"]
    )
    if circles is None:
        return []
    return [(int(cx), int(cy), int(r)) for cx, cy, r in np.round(circles[0]).astype(int)]

# ──────────────────────────────────────────────────────────────────
# CLASSIFICATION  (colour heuristic — replace with ONNX model if available)
# ──────────────────────────────────────────────────────────────────
def classify_bunch(frame: np.ndarray, cx: int, cy: int, r: int) -> tuple[str, float]:
    """
    Sample HSV values inside the detected circle and classify ripeness.
    Tuned to the colour characteristics of this video (outdoor CCTV, daylight).
    Returns (class_name, confidence).
    """
    h_fr, w_fr = frame.shape[:2]
    mask = np.zeros((h_fr, w_fr), dtype=np.uint8)
    cv2.circle(mask, (cx, cy), max(1, r - 3), 255, -1)   # slight inset to avoid edges

    hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
    roi = hsv[mask == 255]
    if len(roi) < 10:
        return "Ripe", 0.50

    mh = float(roi[:, 0].mean())
    ms = float(roi[:, 1].mean())
    mv = float(roi[:, 2].mean())

    # Std for confidence proxy
    sv = float(roi[:, 2].std())

    # Very dark → Unripe
    if mv < 55:
        return "Unripe", min(0.90, 0.60 + (55 - mv) / 55 * 0.3)

    # Very dark + low sat → Rotten
    if mv < 80 and ms < 55:
        return "Rotten", 0.68

    # Bright + very desaturated → Empty bunch (spikelets, no fruits)
    if mv > 170 and ms < 35:
        return "Empty", 0.72

    # Orange-red hue (ripe FFB)  H=0-20 or 155-180
    orange_pix = roi[(roi[:, 0] < 22) | (roi[:, 0] > 155)]
    orange_frac = len(orange_pix) / len(roi)
    if orange_frac > 0.30 and ms > 50:
        return "Ripe", min(0.95, 0.55 + orange_frac)

    # Purple-red (underripe) H=115-160
    purple_pix = roi[(roi[:, 0] > 115) & (roi[:, 0] < 162) & (roi[:, 1] > 45)]
    purple_frac = len(purple_pix) / len(roi)
    if purple_frac > 0.18:
        return "Underripe", min(0.85, 0.55 + purple_frac)

    # Default: treat as Ripe with moderate confidence
    return "Ripe", 0.55

# ──────────────────────────────────────────────────────────────────
# WEIGHT ESTIMATION
# ──────────────────────────────────────────────────────────────────
def estimate_weight(r_px: int, frame_w: int) -> float:
    """
    Bunch radius in pixels → real-world diameter → MPOB weight estimate.
    Assumes camera is ~3-5 m above ramp, scene_width_m calibrated.
    """
    px_per_m   = frame_w / CFG["scene_width_m"]
    diameter_m = (r_px * 2) / px_per_m
    diameter_m = max(0.18, min(0.60, diameter_m))
    t          = (diameter_m - 0.18) / (0.60 - 0.18)
    kg         = CFG["weight_min_kg"] + t * (CFG["weight_max_kg"] - CFG["weight_min_kg"])
    return round(kg, 1)

# ──────────────────────────────────────────────────────────────────
# GRID ID ASSIGNMENT
# ──────────────────────────────────────────────────────────────────
def assign_grid(cx: int, cy: int, frame_w: int, frame_h: int) -> str:
    """Map pixel position to grid cell label e.g. R2C4."""
    col = min(int(cx / frame_w * CFG["grid_cols"]) + 1, CFG["grid_cols"])
    row = min(int(cy / frame_h * CFG["grid_rows"]) + 1, CFG["grid_rows"])
    return f"R{row}C{col}"

# ──────────────────────────────────────────────────────────────────
# SIMPLE IOU-BASED TRACKER
# ──────────────────────────────────────────────────────────────────
def circle_iou(c1, c2) -> float:
    """Approximate IoU of two circles as overlap / union of bounding squares."""
    cx1, cy1, r1 = c1
    cx2, cy2, r2 = c2
    dist = math.hypot(cx1 - cx2, cy1 - cy2)
    if dist > r1 + r2:
        return 0.0
    # Use area of intersection of two circles
    if dist <= abs(r1 - r2):
        smaller = min(r1, r2)
        return (math.pi * smaller**2) / (math.pi * max(r1, r2)**2)
    # Partial intersection
    d = dist
    a = r1**2 * math.acos((d**2 + r1**2 - r2**2) / (2 * d * r1))
    b = r2**2 * math.acos((d**2 + r2**2 - r1**2) / (2 * d * r2))
    c_val = 0.5 * math.sqrt((-d + r1 + r2) * (d + r1 - r2) * (d - r1 + r2) * (d + r1 + r2))
    intersection = a + b - c_val
    union = math.pi * (r1**2 + r2**2) - intersection
    return max(0.0, intersection / max(union, 1))

class BunchTracker:
    def __init__(self):
        self.bunches: dict[int, Bunch] = {}
        self._next_id = 1

    def update(self, detections: list[tuple[int,int,int]],
               frame: np.ndarray, frame_w: int, frame_h: int):
        """
        Greedy distance-based matching. Tracks persist until missed 8 times.
        Confirmed after seen stable_after_n times consecutively.
        """
        used_existing = set()
        used_new      = set()
        matches       = []

        existing_list = list(self.bunches.values())
        # Sort detections by area desc for stable matching
        for det_i, (cx, cy, r) in enumerate(detections):
            best_dist = 60   # max pixel distance to match
            best_id   = None
            for b in existing_list:
                if b.track_id in used_existing:
                    continue
                dist = math.hypot(cx - b.cx, cy - b.cy)
                if dist < best_dist:
                    best_dist = dist
                    best_id   = b.track_id
            if best_id is not None:
                matches.append((best_id, det_i))
                used_existing.add(best_id)
                used_new.add(det_i)

        # Update matched
        for bid, det_i in matches:
            cx, cy, r = detections[det_i]
            b = self.bunches[bid]
            b.cx, b.cy, b.radius = cx, cy, r
            b.stable_cnt += 1
            b._miss_cnt   = 0
            if b.stable_cnt >= CFG["stable_after_n"] and not b.confirmed:
                b.confirmed = True
                b.cls, b.conf = classify_bunch(frame, cx, cy, r)
                b.weight_kg   = estimate_weight(r, frame_w)
                b.grid_id     = assign_grid(cx, cy, frame_w, frame_h)
                margin = int(r * 1.3)
                y0 = max(0, cy - margin); y1 = min(frame_h, cy + margin)
                x0 = max(0, cx - margin); x1 = min(frame_w, cx + margin)
                crop = frame[y0:y1, x0:x1]
                if crop.size > 0:
                    b.thumb = cv2.resize(crop, CFG["thumb_size"])

        # New detections
        for det_i, (cx, cy, r) in enumerate(detections):
            if det_i not in used_new:
                bid = self._next_id; self._next_id += 1
                nb  = Bunch(track_id=bid, cx=cx, cy=cy, radius=r,
                            first_seen=datetime.now().strftime("%H:%M:%S"))
                nb._miss_cnt  = 0
                self.bunches[bid] = nb

        # Increment miss counter for unmatched; prune after 8 misses
        prune = []
        for bid, b in self.bunches.items():
            if bid not in used_existing:
                if not hasattr(b, "_miss_cnt"):
                    b._miss_cnt = 0
                b._miss_cnt += 1
                if b._miss_cnt > 8 and not b.confirmed:
                    prune.append(bid)
        for bid in prune:
            del self.bunches[bid]

    def confirmed_bunches(self) -> list[Bunch]:
        return [b for b in self.bunches.values() if b.confirmed]

# ──────────────────────────────────────────────────────────────────
# DISPLAY / ANNOTATION
# ──────────────────────────────────────────────────────────────────
FONT      = cv2.FONT_HERSHEY_SIMPLEX
FONT_BOLD = cv2.FONT_HERSHEY_DUPLEX

def draw_bunch_label(frame: np.ndarray, b: Bunch):
    """Draw circle + all parameter labels on the frame for one bunch."""
    cx, cy, r = b.cx, b.cy, b.radius
    color      = C_BGR[b.cls]
    h_fr, w_fr = frame.shape[:2]

    # Outer ring
    cv2.circle(frame, (cx, cy), r + 2, (0, 0, 0), 1)          # shadow
    cv2.circle(frame, (cx, cy), r,     color,     2)            # main ring
    cv2.circle(frame, (cx, cy), 3,     (255,255,255), -1)       # centre dot

    # Semi-transparent fill
    overlay = frame.copy()
    cv2.circle(overlay, (cx, cy), r, color, -1)
    cv2.addWeighted(overlay, 0.18, frame, 0.82, 0, frame)
    cv2.circle(frame, (cx, cy), r, color, 2)   # re-draw border after blend

    # Build label block
    label_lines = [
        (f"#{b.track_id}  {b.grid_id}",  1.0,  (255, 255, 255)),
        (f"{b.cls}",                      1.0,  color),
        (f"Conf: {b.conf*100:.0f}%",      0.85, (200, 200, 200)),
        (f"Wt: {b.weight_kg:.1f} kg",     0.85, (220, 240, 220)),
        (f"({b.wt_class})",               0.75, (180, 180, 180)),
        (f"OER: {b.graded_oer:.1f}%",     0.80, (180, 220, 180)),
        (f"RM {b.price_rm:.2f}",          0.90, (80, 255, 120)),
    ]

    # Position label: prefer above, else below
    font_scale = 0.30
    line_h     = 13
    block_h    = len(label_lines) * line_h + 6
    block_w    = 105

    lx = cx - block_w // 2
    ly = cy - r - block_h - 6
    if ly < 5:
        ly = cy + r + 8
    lx = max(2, min(w_fr - block_w - 2, lx))

    # Background pill
    cv2.rectangle(frame, (lx - 2, ly - 2), (lx + block_w, ly + block_h),
                  (20, 20, 20), -1)
    cv2.rectangle(frame, (lx - 2, ly - 2), (lx + block_w, ly + block_h),
                  color, 1)

    for i, (txt, scale_mult, col) in enumerate(label_lines):
        cv2.putText(frame, txt, (lx + 2, ly + 5 + i * line_h),
                    FONT, font_scale * scale_mult, col, 1, cv2.LINE_AA)


def draw_hud(frame: np.ndarray, bunches: list[Bunch],
             fps: float, frame_no: int, paused: bool):
    """Draw the heads-up display panel on the right side."""
    h_fr, w_fr = frame.shape[:2]
    confirmed = bunches

    total_w  = sum(b.weight_kg for b in confirmed)
    total_rm = sum(b.price_rm  for b in confirmed)
    ripe_n   = sum(1 for b in confirmed if b.cls == "Ripe")
    ripe_pct = ripe_n / max(len(confirmed), 1) * 100

    # Counts per class
    cls_cnt = {c: sum(1 for b in confirmed if b.cls == c) for c in CLASSES}

    # HUD panel
    panel_x = w_fr - 185
    cv2.rectangle(frame, (panel_x, 0), (w_fr, h_fr), (15, 15, 15), -1)
    cv2.line(frame, (panel_x, 0), (panel_x, h_fr), (60, 60, 60), 1)

    y = 16
    def hud_text(txt, color=(220,220,220), scale=0.38, bold=False):
        nonlocal y
        f = FONT_BOLD if bold else FONT
        cv2.putText(frame, txt, (panel_x + 6, y), f, scale, color, 1, cv2.LINE_AA)
        y += int(scale * 50 + 2)

    hud_text("MPOB FFB SCANNER", (80, 210, 80), 0.38, bold=True)
    hud_text(datetime.now().strftime("%d/%m/%Y %H:%M:%S"), (160,160,160), 0.30)
    hud_text(f"Frame #{frame_no}  FPS:{fps:.1f}", (140,140,140), 0.30)
    if paused:
        hud_text("[ PAUSED ]", (50, 50, 255), 0.38, bold=True)
    y += 4

    hud_text("─── SUMMARY ───────────", (70,70,70), 0.30)
    hud_text(f"Bunches:  {len(confirmed)}", (220,220,220), 0.36, bold=True)
    hud_text(f"Weight:   {total_w:.1f} kg", (180,240,180), 0.34)
    hud_text(f"Value:    RM {total_rm:.2f}", (100,255,130), 0.36, bold=True)
    hud_text(f"Ripe %:   {ripe_pct:.1f}%", (80,210,80) if ripe_pct>=90 else (255,165,0), 0.34)
    y += 4

    hud_text("─── BY CLASS ───────────", (70,70,70), 0.30)
    for cls in CLASSES:
        n = cls_cnt.get(cls, 0)
        pct = n / max(len(confirmed), 1) * 100
        col = C_BGR[cls]
        hud_text(f"{cls[:5]:5s}  {n:3d}  {pct:5.1f}%", col, 0.32)
    y += 4

    hud_text("─── MPOB LIMITS ────────", (70,70,70), 0.30)
    hud_text("Ripe   ≥90%", (80,210,80), 0.30)
    hud_text("Under  ≤10%", (30,165,255), 0.30)
    hud_text("Unripe  0%", (30,30,200), 0.30)
    hud_text("Rotten  0%", (160,20,180), 0.30)
    hud_text("Empty   0%", (120,120,120), 0.30)
    y += 6

    # Reject warning
    empty_pct = cls_cnt.get("Empty",0) / max(len(confirmed),1) * 100
    if empty_pct > 20:
        hud_text("⚠ REJECT LOAD", (0,0,255), 0.38, bold=True)
        hud_text(f"Empty > 20%: {empty_pct:.1f}%", (0,80,255), 0.32)
    y += 6

    hud_text("─── CONTROLS ───────────", (70,70,70), 0.30)
    hud_text("S = Snapshot", (160,160,160), 0.30)
    hud_text("R = Reset", (160,160,160), 0.30)
    hud_text("SPACE = Pause", (160,160,160), 0.30)
    hud_text("Q = Quit+Export", (160,160,160), 0.30)


def draw_grid_overlay(frame: np.ndarray):
    """Draw faint grid lines."""
    h_fr, w_fr = frame.shape[:2]
    main_w = w_fr - 185
    gcols, grows = CFG["grid_cols"], CFG["grid_rows"]
    cw = main_w // gcols
    ch = h_fr    // grows

    for c in range(1, gcols):
        x = c * cw
        cv2.line(frame, (x, 0), (x, h_fr), (50, 50, 50), 1)
    for r in range(1, grows):
        y = r * ch
        cv2.line(frame, (0, y), (main_w, y), (50, 50, 50), 1)

    # Grid labels top-left corner of each cell
    for r in range(grows):
        for c in range(gcols):
            lbl = f"R{r+1}C{c+1}"
            cv2.putText(frame, lbl, (c*cw + 3, r*ch + 11),
                        FONT, 0.22, (55, 55, 55), 1, cv2.LINE_AA)


# ──────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────────────────────────
def _side(s="thin"): return Side(style=s, color="BDBDBD")
BORD  = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
HFILL = PatternFill("solid", fgColor="1B5E20")
HFONT = Font(bold=True, color="FFFFFF", name="Arial", size=9)
BFONT = Font(name="Arial", size=9)
TFONT = Font(name="Arial", size=14, bold=True, color="1B5E20")
SFONT = Font(name="Arial", size=8, italic=True, color="616161")
CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _w(ws, rc, val, font=None, fill=None, aln=None, bdr=None, fmt=None):
    c = ws[rc]
    c.value = val
    if font: c.font      = font
    if fill: c.fill      = fill
    if aln:  c.alignment = aln
    if bdr:  c.border    = bdr
    if fmt:  c.number_format = fmt


def export_excel(bunches: list[Bunch], snapshot: Optional[np.ndarray],
                 meta: dict) -> str:
    """Build and save a 3-sheet Excel workbook."""
    Path(CFG["output_dir"]).mkdir(exist_ok=True)
    wb = Workbook()
    _sheet_summary(wb, bunches, snapshot, meta)
    _sheet_detail(wb, bunches)
    _sheet_grid(wb, bunches)

    out = str(Path(CFG["output_dir"]) / CFG["excel_name"])
    wb.save(out)
    return out


def _sheet_summary(wb, bunches, snap, meta):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    for col, w in [("A",22),("B",14),("C",14),("D",14),("E",14),("F",14),("G",14),("H",14)]:
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells("A1:H1")
    _w(ws, "A1", "🌴  MPOB Oil Palm FFB — Real-Time CCTV Scan Report", TFONT, aln=CTR)
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:H2")
    _w(ws, "A2",
       f"Scan time: {meta['ts']}   |  Source: {meta['source']}   |  "
       f"Price: RM {CFG['price_rm_per_tonne']:.0f}/tonne   |  "
       f"Grid: {CFG['grid_rows']}×{CFG['grid_cols']}",
       SFONT, aln=LFT)
    ws.row_dimensions[2].height = 13

    # Snapshot image
    img_end = 3
    if snap is not None:
        sw, sh = 420, 260
        thumb  = cv2.cvtColor(cv2.resize(snap, (sw, sh)), cv2.COLOR_BGR2RGB)
        buf    = io.BytesIO(); Image.fromarray(thumb).save(buf, "PNG"); buf.seek(0)
        xl = XLImage(buf); xl.width = sw; xl.height = sh
        ws.add_image(xl, "A4")
        for rr in range(4, 4+20): ws.row_dimensions[rr].height = sh/20
        img_end = 24

    # KPI row
    confirmed = bunches
    total_w   = sum(b.weight_kg for b in confirmed)
    total_rm  = sum(b.price_rm  for b in confirmed)
    ripe_pct  = sum(1 for b in confirmed if b.cls=="Ripe") / max(len(confirmed),1) * 100

    kr = img_end + 1
    for (col, lbl, val, fmt) in [
        ("A","Total Bunches",   len(confirmed),  "0"),
        ("C","Total Weight (kg)", total_w,        "#,##0.0"),
        ("E","Est. Value (RM)",  total_rm,        '"RM"#,##0.00'),
        ("G","Ripe %",          ripe_pct,         '0.0"%"'),
    ]:
        c2 = chr(ord(col)+1)
        ws.merge_cells(f"{col}{kr}:{c2}{kr}")
        ws.merge_cells(f"{col}{kr+1}:{c2}{kr+1}")
        _w(ws, f"{col}{kr}", lbl, SFONT, aln=CTR)
        cc = ws[f"{col}{kr+1}"]
        cc.value = val; cc.number_format = fmt
        cc.font  = Font(name="Arial", size=20, bold=True, color="1B5E20")
        cc.alignment = CTR
    ws.row_dimensions[kr].height   = 13
    ws.row_dimensions[kr+1].height = 34

    # Class breakdown table
    hr = kr + 3
    hdrs = ["Class","Count","% of Load","Weight (kg)","Value (RM)",
            "Basic OER %","Oil Penalty %","Graded OER %"]
    for i, h in enumerate(hdrs):
        _w(ws, f"{get_column_letter(i+1)}{hr}", h, HFONT, HFILL, CTR, BORD)
    ws.row_dimensions[hr].height = 18

    for i, cls in enumerate(CLASSES):
        rr   = hr + 1 + i
        grp  = [b for b in bunches if b.cls == cls]
        cnt  = len(grp)
        wt   = sum(b.weight_kg for b in grp)
        val  = sum(b.price_rm  for b in grp)
        pct  = cnt / max(len(confirmed), 1)
        fill = PatternFill("solid", fgColor=C_HEX[cls])
        vals = [cls, cnt, pct, wt, val, BASIC_OER, OIL_PEN[cls], BASIC_OER - OIL_PEN[cls]*pct*100]
        fmts = ["@","0","0.0%","#,##0.0",'"RM"#,##0.00',"0.0","0.000","0.00"]
        for j,(v,f) in enumerate(zip(vals,fmts)):
            _w(ws, f"{get_column_letter(j+1)}{rr}", v, BFONT, fill, CTR, BORD, f)
        ws.row_dimensions[rr].height = 15

    # Totals
    tot = hr + 1 + len(CLASSES)
    tf  = PatternFill("solid", fgColor="E8F5E9")
    ff  = Font(name="Arial", size=9, bold=True, color="1B5E20")
    for j,(v,f) in enumerate(zip(
        ["TOTAL",len(confirmed),1.0,total_w,total_rm,BASIC_OER,"",""],
        ["@","0","0%","#,##0.0",'"RM"#,##0.00',"0.0","",""]
    )):
        _w(ws, f"{get_column_letter(j+1)}{tot}", v, ff, tf, CTR, BORD, f)
    ws.row_dimensions[tot].height = 16

    # Notes
    nr = tot + 2
    ws.merge_cells(f"A{nr}:H{nr}")
    _w(ws, f"A{nr}",
       "MPOB Quality Limits (Sec 5.2.3): Ripe ≥90%  Underripe ≤10%  Unripe/Rotten/Empty = 0%  "
       " | Reject load if Empty >20% or Dirty >30%",
       Font(name="Arial",size=8,italic=True,color="B71C1C"),
       PatternFill("solid",fgColor="FFF9C4"), LFT)
    ws.row_dimensions[nr].height = 13
    ws.merge_cells(f"A{nr+1}:H{nr+1}")
    _w(ws, f"A{nr+1}",
       "⚠  Prototype AI — results must be verified by a qualified MPOB-licensed grading officer.",
       Font(name="Arial",size=8,italic=True,color="777777"), aln=LFT)


def _sheet_detail(wb, bunches):
    ws = wb.create_sheet("Cell Detail")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:Q1")
    _w(ws, "A1", "Cell-by-Cell Detail — One row per detected bunch", TFONT, aln=CTR)
    ws.row_dimensions[1].height = 22

    IW, IH = CFG["thumb_size"]
    hdrs = ["#","Track ID","Grid ID","Class","Confidence %",
            "MPOB Weight Class","Est. Weight (kg)","Est. Price (RM)",
            "Basic OER %","Oil Penalty %","Graded OER %","Graded KER %",
            "Radius (px)","First Seen","Scan Date","Status","Bunch Photo"]
    for i, h in enumerate(hdrs):
        _w(ws, f"{get_column_letter(i+1)}2", h, HFONT, HFILL, CTR, BORD)
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "A3"

    now_str = datetime.now().strftime("%Y-%m-%d")
    for i, b in enumerate(bunches):
        rr   = i + 3
        fill = PatternFill("solid", fgColor=C_HEX[b.cls])
        status = "✅ Good" if b.cls=="Ripe" else ("⚠️ Penalty" if b.cls in ("Underripe",) else "❌ Poor")
        vals = [i+1, b.track_id, b.grid_id, b.cls, b.conf*100,
                b.wt_class, b.weight_kg, b.price_rm,
                BASIC_OER, b.oil_pen, b.graded_oer, b.graded_ker if hasattr(b,'graded_ker') else BASIC_KER,
                b.radius, b.first_seen, now_str, status, ""]
        fmts = ["0","0","@","@","0.0",
                "@","#,##0.0",'"RM"#,##0.00',
                "0.0","0.0000","0.00","0.0",
                "0","@","@","@","@"]
        for j,(v,f) in enumerate(zip(vals,fmts)):
            _w(ws, f"{get_column_letter(j+1)}{rr}", v, BFONT, fill, CTR, BORD, f)

        # Embed bunch thumbnail
        if b.thumb is not None:
            thumb_bgr = b.thumb
            thumb_rgb = cv2.cvtColor(thumb_bgr, cv2.COLOR_BGR2RGB)
            buf = io.BytesIO(); Image.fromarray(thumb_rgb).save(buf,"JPEG",quality=85); buf.seek(0)
            xi = XLImage(buf); xi.width = IW; xi.height = IH
            ws.add_image(xi, f"{get_column_letter(len(hdrs))}{rr}")
        ws.row_dimensions[rr].height = IH * 0.75

    widths = [4,8,8,12,12,14,14,14,10,11,11,10,10,9,12,10,17]
    for i,w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w


def _sheet_grid(wb, bunches):
    """Visual grid map matching the ramp layout."""
    ws = wb.create_sheet("Grid Map")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:Z1")
    _w(ws, "A1", "Grid Map — Bunch position on ramp floor", TFONT, aln=CTR)
    ws.row_dimensions[1].height = 26

    nrows, ncols = CFG["grid_rows"], CFG["grid_cols"]
    IW, IH = CFG["thumb_size"]
    CP, RP = 2, 9     # Excel columns per cell, rows per cell
    SR, SC = 3, 2     # start row, start col (1-indexed)

    # Col headers
    for c in range(ncols):
        ec  = get_column_letter(SC + c*CP)
        ec2 = get_column_letter(SC + c*CP + 1)
        ws.merge_cells(f"{ec}2:{ec2}2")
        _w(ws, f"{ec}2", f"Col {c+1}", HFONT, HFILL, CTR)
        ws.column_dimensions[ec].width  = IW / 7.5
        ws.column_dimensions[ec2].width = IW / 7.5
    ws.column_dimensions["A"].width = 5

    # Build a map: grid_id → list of bunches (could be >1 if crowded)
    grid_map: dict[str, list[Bunch]] = defaultdict(list)
    for b in bunches:
        grid_map[b.grid_id].append(b)

    for r in range(nrows):
        base = SR + r * RP
        ws.merge_cells(f"A{base}:A{base+RP-1}")
        _w(ws, f"A{base}", f"Row {r+1}", HFONT, HFILL,
           Alignment(horizontal="center", vertical="center", text_rotation=90))
        for sub in range(RP):
            ws.row_dimensions[base+sub].height = IH / RP * 1.1

        for c in range(ncols):
            gid  = f"R{r+1}C{c+1}"
            cell_bunches = grid_map.get(gid, [])
            ec   = get_column_letter(SC + c*CP)
            ec2  = get_column_letter(SC + c*CP + 1)

            if not cell_bunches:
                # Empty grid cell
                fill = PatternFill("solid", fgColor="F5F5F5")
                ws.merge_cells(f"{ec}{base}:{ec2}{base+RP-1}")
                cc = ws[f"{ec}{base}"]
                cc.fill = fill
                cc.border = BORD
                cc.value = gid
                cc.font  = Font(name="Arial", size=7, color="BBBBBB")
                cc.alignment = CTR
                continue

            # Use first (or best-confidence) bunch
            b    = max(cell_bunches, key=lambda x: x.conf)
            fill = PatternFill("solid", fgColor=C_HEX[b.cls])

            # Colour all rows in this cell
            for sub in range(RP - 2):
                for co in range(CP):
                    cell = ws.cell(row=base+sub, column=SC + c*CP + co)
                    cell.fill   = fill
                    cell.border = BORD

            # Embed thumbnail if available
            if b.thumb is not None:
                thumb_rgb = cv2.cvtColor(b.thumb, cv2.COLOR_BGR2RGB)
                buf = io.BytesIO(); Image.fromarray(thumb_rgb).save(buf,"JPEG",quality=85); buf.seek(0)
                xi = XLImage(buf); xi.width = IW; xi.height = IH - 24
                ws.add_image(xi, f"{ec}{base}")

            # Caption — write BEFORE merging to avoid MergedCell read-only error
            cap_row  = base + RP - 2
            cap_cell = ws.cell(row=cap_row, column=SC + c*CP)
            cap_cell.value = (f"{gid}  {b.cls}\n"
                              f"{b.weight_kg:.1f}kg | RM{b.price_rm:.2f}\n"
                              f"Conf:{b.conf*100:.0f}%")
            cap_cell.font      = Font(name="Arial", size=7, bold=True)
            cap_cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
            cap_cell.fill   = fill
            cap_cell.border = BORD
            ws.merge_cells(f"{ec}{cap_row}:{ec2}{cap_row+1}")


# ──────────────────────────────────────────────────────────────────
# MAIN  —  real-time processing loop
# ──────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="FFB Real-Time CCTV Scanner")
    ap.add_argument("--source", default=None,
                    help="Video file path, RTSP URL, or camera index (default: 0)")
    ap.add_argument("--scene-width",  type=float, default=CFG["scene_width_m"])
    ap.add_argument("--price",        type=float, default=CFG["price_rm_per_tonne"])
    ap.add_argument("--grid-rows",    type=int,   default=CFG["grid_rows"])
    ap.add_argument("--grid-cols",    type=int,   default=CFG["grid_cols"])
    ap.add_argument("--export-frames",action="store_true",
                    help="Save annotated frames to output dir")
    ap.add_argument("--no-display",   action="store_true",
                    help="Run headless (for server/CI use)")
    args = ap.parse_args()

    CFG["scene_width_m"]     = args.scene_width
    CFG["price_rm_per_tonne"]= args.price
    CFG["grid_rows"]         = args.grid_rows
    CFG["grid_cols"]         = args.grid_cols

    Path(CFG["output_dir"]).mkdir(exist_ok=True)

    # Open source
    src = args.source
    if src is None:
        src = 0
    elif src.isdigit():
        src = int(src)

    cap = cv2.VideoCapture(src)
    if not cap.isOpened():
        sys.exit(f"❌  Cannot open source: {src}")

    fps_src    = cap.get(cv2.CAP_PROP_FPS) or 25
    frame_w    = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    frame_h    = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    total_fr   = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    # Add HUD panel width
    HUD_W   = 185
    disp_w  = int((frame_w + HUD_W) * CFG["display_scale"])
    disp_h  = int(frame_h * CFG["display_scale"])

    print(f"✅  Source: {src}  |  {frame_w}×{frame_h} @ {fps_src:.0f}fps")
    print(f"   Grid: {CFG['grid_rows']}×{CFG['grid_cols']}  "
          f"Scene width: {CFG['scene_width_m']}m  "
          f"Price: RM{CFG['price_rm_per_tonne']}/tonne")
    print("   Press Q to quit and export  |  S = snapshot  |  SPACE = pause  |  R = reset")

    tracker    = BunchTracker()
    frame_no   = 0
    fps_disp   = 0.0
    t_prev     = time.time()
    paused     = False
    last_snap  = None   # annotated frame for Excel

    if not args.no_display:
        cv2.namedWindow("FFB CCTV Scanner", cv2.WINDOW_NORMAL)
        cv2.resizeWindow("FFB CCTV Scanner", disp_w, disp_h)

    while True:
        if not paused:
            ok, raw_frame = cap.read()
            if not ok:
                print("\n⏹  End of source — exporting report…")
                break
            frame_no += 1
        else:
            # Keep displaying last frame
            key = cv2.waitKey(30) & 0xFF
            if key == ord(' '): paused = False
            elif key == ord('q'): break
            continue

        # ── Detection every N frames ─────────────────
        if frame_no % CFG["process_every_n"] == 0:
            detections = detect_bunches(raw_frame)
            tracker.update(detections, raw_frame, frame_w, frame_h)

        # ── Build display frame ───────────────────────
        disp = raw_frame.copy()

        # Extend canvas for HUD
        hud_panel = np.zeros((frame_h, HUD_W, 3), dtype=np.uint8)
        disp_full = np.hstack([disp, hud_panel])

        draw_grid_overlay(disp_full)

        confirmed = tracker.confirmed_bunches()
        for b in confirmed:
            draw_bunch_label(disp_full, b)

        # FPS counter
        t_now   = time.time()
        fps_disp = 0.9 * fps_disp + 0.1 * (1.0 / max(t_now - t_prev, 1e-6))
        t_prev  = t_now

        draw_hud(disp_full, confirmed, fps_disp, frame_no, paused)

        last_snap = disp_full.copy()

        if args.export_frames and frame_no % 25 == 0:
            cv2.imwrite(f"{CFG['output_dir']}/frame_{frame_no:06d}.jpg", disp_full)

        if not args.no_display:
            cv2.imshow("FFB CCTV Scanner", disp_full)
            key = cv2.waitKey(1) & 0xFF

            if key == ord('q'):
                print("\n⏹  Q pressed — exporting report…")
                break
            elif key == ord('s'):
                # Manual snapshot
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                snap_path = f"{CFG['output_dir']}/snapshot_{ts}.jpg"
                cv2.imwrite(snap_path, disp_full)
                print(f"📸  Snapshot saved: {snap_path}")
            elif key == ord(' '):
                paused = not paused
                print("⏸  Paused" if paused else "▶  Resumed")
            elif key == ord('r'):
                tracker = BunchTracker()
                print("🔄  Tracker reset")

        # Progress (for video files)
        if total_fr > 0 and frame_no % 50 == 0:
            pct = frame_no / total_fr * 100
            print(f"  [{pct:5.1f}%] frame {frame_no}/{total_fr}  "
                  f"bunches: {len(confirmed)}  "
                  f"weight: {sum(b.weight_kg for b in confirmed):.1f}kg  "
                  f"value: RM{sum(b.price_rm for b in confirmed):.2f}",
                  end="\r")

    cap.release()
    if not args.no_display:
        cv2.destroyAllWindows()

    # ── Export ───────────────────────────────────────
    confirmed = tracker.confirmed_bunches()
    total_w   = sum(b.weight_kg for b in confirmed)
    total_rm  = sum(b.price_rm  for b in confirmed)

    print(f"\n\n{'='*58}")
    print(f"  FINAL SCAN SUMMARY")
    print(f"{'='*58}")
    print(f"  Bunches confirmed  : {len(confirmed)}")
    print(f"  Total weight       : {total_w:.1f} kg")
    print(f"  Estimated value    : RM {total_rm:.2f}")
    for cls in CLASSES:
        n = sum(1 for b in confirmed if b.cls==cls)
        if n:
            print(f"  {cls:<12}: {n:4d}  ({n/max(len(confirmed),1)*100:.1f}%)")
    print(f"{'='*58}")

    if confirmed:
        xls_path = export_excel(
            confirmed, last_snap,
            {"ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
             "source": str(src)}
        )
        print(f"\n✅  Excel report saved : {xls_path}")
    else:
        print("⚠  No confirmed bunches — no Excel exported.")


if __name__ == "__main__":
    main()
